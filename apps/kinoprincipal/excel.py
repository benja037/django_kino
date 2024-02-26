from openpyxl import load_workbook
from apps.kinoprincipal.models import Kinodb, Rekinodb, Chanchitodb, Combodb, Chao1db, Chao2db, Chao3db, Archivosxl
import datetime
import requests
from bs4 import BeautifulSoup



def Importar_datos(archivo):
    # Leer Excel
    wb = load_workbook(filename=archivo)
    ws = wb["Kino_Int"]

    ultima_fila = ws.max_row - 3
    #print(ultima_fila)
    # Parametros Kino
    rango_casillas_kino = "C808:P" + str(ultima_fila)
    int_inicio_kino = 808
    columna_ganadores_kino = "LD"

    # Parametros Rekino
    rango_casillas_rekino = "AD872:AQ" + str(ultima_fila)
    int_inicio_rekino = 872
    columna_ganadores_rekino = "LF"

    # Parametros Chanchito
    rango_casillas_chanchito = "AT1645:BG" + str(ultima_fila)
    int_inicio_chanchito = 1645
    columna_ganadores_chanchito = "LG"

    # Parametros Combo
    rango_casillas_combo = "BX1645:CK" + str(ultima_fila)
    int_inicio_combo = 1645
    columna_ganadores_combo = "LI"

    # Parametros Chao Jefe 1
    rango_casillas_chao1 = "JK2371:JX" + str(ultima_fila)
    int_inicio_chao1 = 2371
    columna_ganadores_chao1 = "LM"

    # Parametros Chao Jefe 2
    rango_casillas_chao2 = "JZ2371:KM" + str(ultima_fila)
    int_inicio_chao2 = 2371
    columna_ganadores_chao2 = "LN"

    # Parametros Chao Jefe 3
    rango_casillas_chao3 = "KO2774:LB" + str(ultima_fila)
    int_inicio_chao3 = 2774
    columna_ganadores_chao3 = "LO"

    excel_a_modelo(ws, Kinodb, rango_casillas_kino,
                   int_inicio_kino, columna_ganadores_kino)
    excel_a_modelo(ws, Rekinodb, rango_casillas_rekino,
                   int_inicio_rekino, columna_ganadores_rekino)
    excel_a_modelo(ws, Chanchitodb, rango_casillas_chanchito,
                   int_inicio_chanchito, columna_ganadores_chanchito)
    excel_a_modelo(ws, Combodb, rango_casillas_combo,
                   int_inicio_combo, columna_ganadores_combo)
    excel_a_modelo(ws, Chao1db, rango_casillas_chao1,
                   int_inicio_chao1, columna_ganadores_chao1)
    excel_a_modelo(ws, Chao2db, rango_casillas_chao2,
                   int_inicio_chao2, columna_ganadores_chao2)
    excel_a_modelo(ws, Chao3db, rango_casillas_chao3,
                   int_inicio_chao3, columna_ganadores_chao3)


def excel_a_modelo(ws, Modelo, rango_casillas, int_inicio, columna_ganadores):

    lista_total = []
    contador = int_inicio
    for row in ws[rango_casillas]:
        str_contador = str(contador)
        sorteo = ws["A" + str_contador].value
        fecha = ws["B" + str_contador].value

        numero1 = int(row[0].value)
        numero2 = int(row[1].value)
        numero3 = int(row[2].value)
        numero4 = int(row[3].value)
        numero5 = int(row[4].value)
        numero6 = int(row[5].value)
        numero7 = int(row[6].value)
        numero8 = int(row[7].value)
        numero9 = int(row[8].value)
        numero10 = int(row[9].value)
        numero11 = int(row[10].value)
        numero12 = int(row[11].value)
        numero13 = int(row[12].value)
        numero14 = int(row[13].value)
        ganador = int(ws[columna_ganadores + str_contador].value)
        tipo_ingreso = "Excel"

        elemento = Modelo(id_sorteo=sorteo, fecha=fecha, number1=numero1, number2=numero2, number3=numero3, number4=numero4, number5=numero5, number6=numero6,
                          number7=numero7, number8=numero8, number9=numero9, number10=numero10, number11=numero11, number12=numero12, number13=numero13, number14=numero14, num_ganadores=ganador, tipo_ingreso=tipo_ingreso)
        lista_total.append(elemento)
        contador = contador + 1

    Modelo.objects.bulk_create(lista_total)




def fetch_kino_results(num_sorteo):
    #print("HOLAAA")
    fecha_hoy = datetime.date.today()
    fecha_ayer = fecha_hoy - datetime.timedelta(days=1)
    resp = requests.get('https://chileresultados.com/kino/sorteos/' + str(num_sorteo))
    #print("LINK",'https://chileresultados.com/kino/sorteos/' + str(num_sorteo))
    numeros = []
    texto = resp.text
    soup = BeautifulSoup(texto, "lxml")
    #print(texto)
    # Filtro a html encuentra números de los 7 sorteos principales
    soup_numeros = soup.find_all(
        "span", class_="badge rounded-pill bg-warning text-dark")
    for elemento in soup_numeros:
        numerito = elemento.get_text()
        numeros.append(int(numerito))

    # Filtro a html encuentra data de la imagen donde el titulo contiene el numero de sorteo
    soup_sorteo = soup.find_all(class_="img-fluid")
    #print("SOUPPPP",soup_sorteo)
    titulo = soup_sorteo[1].get('title')
    #print("SSSSSSSSSSSSS",titulo.split(" ")[2])
    sorteo = int(titulo.split(" ")[2])

    resultados_kino = numeros[0:14]
    resultados_rekino = numeros[14:28]
    resultados_chanchito = numeros[28:42]
    resultados_combo = numeros[42:56]
    resultados_chao1 = numeros[56:70]
    resultados_chao2 = numeros[70:84]
    resultados_chao3 = numeros[84:98]
    print("kino")
    pasar_modelo(Kinodb, resultados_kino, sorteo, fecha_ayer)
    print("rekino")
    pasar_modelo(Rekinodb, resultados_rekino, sorteo, fecha_ayer)
    print("chanchito")
    pasar_modelo(Chanchitodb, resultados_chanchito, sorteo, fecha_ayer)
    print("combo")
    pasar_modelo(Combodb, resultados_combo, sorteo, fecha_ayer)
    print("chao1")
    pasar_modelo(Chao1db, resultados_chao1, sorteo, fecha_ayer)
    print("chao2")
    pasar_modelo(Chao2db, resultados_chao2, sorteo, fecha_ayer)
    print("chao3")
    pasar_modelo(Chao3db, resultados_chao3, sorteo, fecha_ayer)


def pasar_modelo(tipo_sorteo_db, tipo_numeros, numero_sorteo, fecha):
    number1 = tipo_numeros[0]
    number2 = tipo_numeros[1]
    number3 = tipo_numeros[2]
    number4 = tipo_numeros[3]
    number5 = tipo_numeros[4]
    number6 = tipo_numeros[5]
    number7 = tipo_numeros[6]
    number8 = tipo_numeros[7]
    number9 = tipo_numeros[8]
    number10 = tipo_numeros[9]
    number11 = tipo_numeros[10]
    number12 = tipo_numeros[11]
    number13 = tipo_numeros[12]
    number14 = tipo_numeros[13]
    tipo_ingreso = "semi-auto"

    print(tipo_sorteo_db, fecha, numero_sorteo, number1, " ", number2, " ", number3, " ", number4, " ", number5, " ", number6, " ", number7, " ",
          number8, " ", number9, " ", number10, " ", number11, " ", number12, " ", number13, " ", number14, " ", tipo_ingreso)

    tipo_sorteo_db.objects.create(id_sorteo=numero_sorteo, fecha=fecha, number1=number1, number2=number2, number3=number3, number4=number4, number5=number5, number6=number6,
                                  number7=number7, number8=number8, number9=number9, number10=number10, number11=number11, number12=number12, number13=number13, number14=number14, tipo_ingreso=tipo_ingreso)
